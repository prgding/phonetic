from openpyxl import load_workbook
from multiprocessing import Pool
import re
import urllib.request as req


def search(e, url):
    r = req.Request(url + e)
    s = req.urlopen(r)
    page = s.read().decode('utf-8')
    # print(page)
    try:
        v = re.search('>美\s+<span class="phonetic">([\S\W\]]+?)<', page)
        # print(e + "\t" + v.group(1))
        return v.group(1)

    except(AttributeError):
        print("检查单词拼写" + e)


if __name__ == '__main__':
    wb = load_workbook(filename='/Users/dingshuai/Dev/Code/Python/phonetic/words.xlsx')
    sheet1 = wb['Sheet1']
    raw_lines=[]
    raw_words = []
    res_pool = []

    for i in range(2, sheet1.max_row + 1):
        words_line = sheet1.cell(row=i, column=1).value
        phonetic_line = sheet1.cell(row=i, column=6).value
        if words_line is not None and phonetic_line is None:
            # 打印出行数
            raw_lines.append(i)
            raw_words.append(sheet1.cell(row=i, column=1).value)
    # print(raw_words)

    url = 'http://dict.youdao.com/w/eng/'
    p = Pool(20)
    for word in raw_words:
        res = p.apply_async(search, args=(word, url))
        res_pool.append(res)
    for index, phonetic in enumerate(res_pool):
        print(index, phonetic.get())
        sheet1.cell(row=min(raw_lines) + index, column=6).value = phonetic.get()

    wb.save('/Users/dingshuai/Dev/Code/Python/phonetic/words.xlsx')
